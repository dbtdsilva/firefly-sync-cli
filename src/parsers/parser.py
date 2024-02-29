import re
import openpyxl
import warnings
import csv
from abc import ABC, abstractmethod
from typing import List

from .types.parsed_transaction import ParsedTransaction


class Parser(ABC):

    @staticmethod
    @abstractmethod
    def parse(file) -> List[ParsedTransaction]:
        pass

    @staticmethod
    def read_table_from_excel(file, start_text, capitalize=False):
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        wb = openpyxl.load_workbook(file)
        sheet = wb.active

        # Find the starting row and column of the table
        start_row = None
        start_col = None
        for row in sheet.iter_rows():
            if row[0].value == start_text:
                start_row = row[0].row
                start_col = row[0].column
                break

        if not start_row or not start_col:
            raise ValueError("Could not find table starting with 'Execution date'")

        # Extract headers
        headers = [cell.value for cell in sheet[start_row]]

        # Extract data and convert to dictionaries
        data = []
        for row in sheet.iter_rows():
            if row[0].row <= start_row:
                continue
            data_row = {Parser.__clean_value(headers[i], capitalize=capitalize):
                        Parser.__clean_value(cell.value, capitalize=capitalize)
                        for i, cell in enumerate(row)}
            data.append(data_row)

        return data

    @staticmethod
    def read_table_from_csv(file, delimiter=',', encoding='utf-8', capitalize=False):
        data = []
        with open(file=file, newline='', encoding=encoding) as csvfile:
            csvreader = csv.reader(csvfile, delimiter=delimiter)
            headers = next(csvreader)
            for row in csvreader:
                parsed_row = {}
                for header, value in zip(headers, row):
                    header_clean = Parser.__clean_value(header, capitalize=capitalize)
                    value_clean = Parser.__clean_value(value, capitalize=capitalize)
                    if header_clean not in parsed_row:
                        parsed_row[header_clean] = value_clean
                data.append(parsed_row)
        return data

    @staticmethod
    def __clean_value(value, capitalize=False):
        if not isinstance(value, str):
            return value

        if capitalize:
            value = value.capitalize()
        value = re.sub(' +', ' ', value.strip())
        if value == '':
            return None
        return value
