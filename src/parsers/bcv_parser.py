import csv
import logging
from typing import List
from datetime import datetime
import openpyxl
import numbers

from ..firefly_api.models.transaction import Transaction, TransactionType
from .parser import Parser

class BcvParser(Parser):

    @staticmethod
    def read_table_from_excel(file_path, start_text):
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        # Find the starting row and column of the table
        start_row = None
        start_col = None
        for row in sheet.iter_rows():
            if row[0].value == start_text:
                start_row = row[0].row
                start_col = row[0].column
                break

        if not start_row or not start_col:
            raise ValueError("Could not find table starting with 'Execution date'")

        # Extract headers
        headers = [cell.value for cell in sheet[start_row]]

        # Extract data and convert to dictionaries
        data = []
        for row in sheet.iter_rows():
            if row[0].row <= start_row:
                continue
            data_row = {headers[i]: cell.value for i, cell in enumerate(row)}
            data.append(data_row)

        return data

    @staticmethod
    def parse(file: str) -> List[Transaction]:
        data = BcvParser.read_table_from_excel(file_path=file, start_text='Execution date')
        # Now data contains the parsed CSV data
        transactions = []
        for row in data:
            transaction_date = datetime.strptime(row['Execution date'], '%d.%m.%Y')
            description = row['Transactions']
            debit = row['Debit'] if isinstance(row['Debit'], numbers.Number) else None
            credit = row['Credit'] if isinstance(row['Credit'], numbers.Number) else None

            if (debit is not None and credit is not None) or \
                (debit is None and credit is None):
                raise Exception(f'Invalid debit / credit in {file}')
            elif debit is not None:
                amount = debit
                transaction_type = TransactionType.WITHDRAWAL
                source_name = 'BCV Prive Privilege'
                destination_name = 'Unidentified'
            elif credit is not None:
                amount = credit
                transaction_type = TransactionType.DEPOSIT
                source_name = 'Unidentified'
                destination_name = 'BCV Prive Privilege'

            transactions.append(Transaction(type=transaction_type, source_name=source_name, 
                                            destination_name=destination_name,
                                            date=transaction_date, amount=amount, 
                                            description=description, currency_code='CHF', 
                                            reconciled=True))
        return transactions